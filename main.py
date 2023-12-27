from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl

# Set up the WebDriver (you need to specify the path to your WebDriver)
driver = webdriver.Chrome()

# Load the web page
driver.get('https://www.educationportal.mp.gov.in/public/search/searchemployee.aspx')

# Open the Excel file and get combinations
wb = openpyxl.load_workbook(r'G:\Customers\Freelancer\Scraping Text Data\combinations.xlsx')
ws = wb.active

# Create a new Excel workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Define the header row
header = ['Unique ID', 'Name', 'DOB', 'Category', 'Designation', 'Subject', 'Posted At', 'Posted At', 'District', 'Mobile', 'DoJ', 'Joined As', 'Payment Authority']

# Write the header row to the worksheet
worksheet.append(header)

def get_next_page_button(driver):
    try:
        return WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//tr[contains(@class, "ui-state-default")]/td/table/tbody/tr/td/a[text()="2"]'))
        )
    except:
        return None

# Loop through the rows in the input Excel file
for row in ws.iter_rows(min_row=3201, max_row=3202, min_col=1, max_col=4):
    first_name, last_name = row[0].value, row[1].value

    designation_type = "Teaching"
    district = "- ALL District -"

    # Wait for the first name input element to be clickable
    first_name_input = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_txtFName')))
    first_name_input.clear()
    first_name_input.send_keys(first_name)

    # Wait for the last name input element to be clickable
    last_name_input = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_txtLName')))
    last_name_input.clear()
    last_name_input.send_keys(last_name)

    # Wait for the district dropdown to be clickable
    district_dropdown = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_ddlDistrict')))
    district_dropdown.send_keys(district)

    # Wait for the designation type dropdown to be clickable
    designation_dropdown = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_ddlDG_Type')))
    designation_dropdown.send_keys(designation_type)

    # Click on the search button
    search_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_btn_Search')))
    search_button.click()

    # Wait for the table to be present (add an appropriate timeout if needed)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_GVEMP_Name')))

    # Extract data using BeautifulSoup
    soup = BeautifulSoup(driver.page_source, 'html.parser')

    # Find the table element with the specific ID
    table = soup.find('table', {'id': 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_GVEMP_Name'})

    # Initialize a list to store the data
    employee_data = []

    # Extract rows (skip the header row)
    rows = table.find_all('tr')[1:]

    # Iterate through rows and extract data
    for row in rows:
        columns = row.find_all('td')

        # Define default values in case a column is missing in the row
        unique_id = name = dob = category = designation = subject = posted_at = posted_at2 = district = mobile = doj = joined_as = payment_authority = ''

        if len(columns) > 0:
            # Check if an <a> element is found in the first column
            a_element = columns[0].find('a')
            if a_element:
                unique_id = a_element.text.strip()
        if len(columns) > 2:
            name = columns[2].text.strip()
        if len(columns) > 3:
            dob = columns[3].text.strip()
        if len(columns) > 4:
            category = columns[4].text.strip()
        if len(columns) > 10:
            designation = columns[10].text.strip()
        if len(columns) > 6:
            subject = columns[6].text.strip()
        if len(columns) > 7:
            posted_at = columns[7].text.strip()
        if len(columns) > 8:
            posted_at2 = columns[8].text.strip()
        if len(columns) > 9:
            district = columns[9].text.strip()
        if len(columns) > 10:
            mobile = columns[10].text.strip()
        if len(columns) > 11:
            doj = columns[11].text.strip()
        if len(columns) > 12:
            joined_as = columns[12].text.strip()
        if len(columns) > 13:
            payment_authority = columns[13].text.strip()

        # Append the extracted data to the list
        employee_data.append([unique_id, name, dob, category, designation, subject, posted_at, posted_at2, district, mobile, doj, joined_as, payment_authority])

    # Save data to Excel file after all rows are processed
    for data_row in employee_data:
        worksheet.append(data_row)
        print(data_row)

    while True:
        # next_page_button = get_next_page_button(driver)
        next_page_button = driver.find_element(By.XPATH, '//tr[@class="ui-state-default"]//a[contains(@href, "__doPostBack") and contains(., "2")]')
        print(next_page_button, 'im the button')
        # next_page_button.click()


        if next_page_button:
            print("Clicking the next Page button")
            next_page_button.click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'ctl00_ctl00_RMCPH_ContentPlaceHolder1_GVEMP_Name')))
        else:
            print('there is no button')
            break


# Save the Excel workbook to a file
excel_filename = 'Employee_details6.xlsx'
workbook.save(excel_filename)

# Close the browser
driver.quit()

print(f'Data has been scraped and saved to {excel_filename}')
